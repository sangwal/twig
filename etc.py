#!/usr/bin/env python3
"""
    etc.py - excel to text converter and vice versa
    ---------------------------------------------------
    Convert Excel (.xlsx) files to a structured text format and vice versa.
    Usage:
        To convert Excel to text:
            python etc.py text source.xlsx destination.txt [--summary]
        To convert text back to Excel:
            python etc.py excel source.txt destination.xlsx [--summary]

    The structured text format represents each sheet and cell with type annotations.
"""
import argparse
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import re
import sys


def excel_to_text(excel_path, text_path, summary=False):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    sheet_count = 0
    cell_count = 0

    with open(text_path, "w", encoding="utf-8") as f:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            sheet_count += 1
            f.write(f"[Sheet: {sheet}]\n")
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue  # skip empty
                    cell_count += 1

                    # Determine cell type and format
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        dtype, val_str = "formula", cell.value
                    else:
                        value = cell.value
                        if isinstance(value, bool):
                            dtype, val_str = "bool", str(value)
                        elif isinstance(value, int):
                            dtype, val_str = "int", str(value)
                        elif isinstance(value, float):
                            dtype, val_str = "float", repr(value)
                        elif isinstance(value, datetime):
                            dtype, val_str = "date", value.isoformat()
                        else:
                            dtype, val_str = "str", str(value).replace("\n", "\\n")

                    f.write(f"{cell.coordinate}: [{dtype}] {val_str}\n")
            f.write("---\n")

    print(f"✅ Workbook '{excel_path}' converted to '{text_path}'.")
    if summary:
        print(f"📄 Sheets processed: {sheet_count}")
        print(f"🔢 Non-empty cells written: {cell_count}")


def text_to_excel(text_path, excel_path, summary=False):
    wb = Workbook()
    ws = None
    pattern = re.compile(r"^([A-Z]+\d+): \[(\w+)\] (.*)$")

    sheet_count = 0
    cell_count = 0

    with open(text_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith("[Sheet: "):
                sheet_name = line[len("[Sheet: "):-1]
                ws = wb.create_sheet(title=sheet_name)
                sheet_count += 1
            elif line == "---":
                continue
            else:
                m = pattern.match(line)
                if not m or ws is None:
                    continue
                coord, dtype, val_str = m.groups()

                if dtype == "int":
                    val = int(val_str)
                elif dtype == "float":
                    val = float(val_str)
                elif dtype == "bool":
                    val = (val_str == "True")
                elif dtype == "date":
                    val = datetime.fromisoformat(val_str)
                elif dtype == "formula":
                    val = val_str
                else:
                    val = val_str.replace("\\n", "\n")

                ws[coord] = val
                cell_count += 1

    # Remove the default empty sheet if unused
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(excel_path)
    print(f"✅ Text file '{text_path}' converted back to '{excel_path}'.")
    if summary:
        print(f"📄 Sheets created: {sheet_count}")
        print(f"🔢 Cells written: {cell_count}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel (.xlsx) ↔ structured text format."
    )
    parser.add_argument(
        "mode", choices=["text", "excel"],
        help="Conversion mode: 'text' = Excel→Text, 'excel' = Text→Excel"
    )
    parser.add_argument("source", help="Source file path")
    parser.add_argument("destination", help="Destination file path")
    parser.add_argument(
        "--summary", action="store_true",
        help="Show number of sheets and cells processed"
    )

    args = parser.parse_args()

    if args.mode == "text":
        excel_to_text(args.source, args.destination, args.summary)
    elif args.mode == "excel":
        text_to_excel(args.source, args.destination, args.summary)
    else:
        print("Invalid mode. Use 'text' or 'excel'.")
        sys.exit(1)


if __name__ == "__main__":
    main()
