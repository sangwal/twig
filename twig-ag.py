#!/usr/bin/env python3
"""
    twig-ag.py -- TeacherWIse [timetable] Generator

    A python script to generate Teacherwise timetable from Classwise
    timetable and individual classwise sheets for all classes.

    Refactored to use modern Python classes and dataclasses.
"""
from __future__ import annotations

import argparse
import configparser
import logging
import re
import sys
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

__version__ = '251228'  # Refactored version

# Global Constants (defaults)
DEFAULT_MAX_PERIODS = 8
DEFAULT_SEPARATOR = "\n"

@dataclass
class TimetableEntry:
    period: int
    class_name: str
    days: str
    subject: str
    teacher: str

    @property
    def expanded_days(self) -> List[int]:
        return DayUtils.expand_days(self.days)


class DayUtils:
    """Utility class for handling day string parsing and compression."""
    
    @staticmethod
    def expand_days(days: str) -> List[int]:
        """Expand a day string like '1-2, 4' into [1, 2, 4]."""
        ret: List[int] = []
        for group in re.split(r',\s*', days):
            if '-' in group:
                try:
                    parts = group.split('-')
                    if len(parts) != 2:
                        continue
                    start_day, end_day = int(parts[0]), int(parts[1])
                    if end_day < start_day:
                        start_day, end_day = end_day, start_day
                    ret.extend(range(start_day, end_day + 1))
                except ValueError:
                    continue
            else:
                try:
                    ret.append(int(group))
                except ValueError:
                    continue
        return ret

    @staticmethod
    def compress_days(days: List[int]) -> str:
        """Compress list of days [1, 2, 4] into '1-2, 4'."""
        if not days:
            return ""
        days = sorted(set(days))
        ranges = []
        start = prev = days[0]
        for day in days[1:]:
            if day == prev + 1:
                prev = day
            else:
                ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
                start = prev = day
        ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
        return ", ".join(ranges)

    @staticmethod
    def count_days(days: str) -> int:
        return len(set(DayUtils.expand_days(days)))


class ConfigManager:
    """Manages application configuration."""
    _instance = None
    _config: Dict[str, Any] = {}

    def __new__(cls, *args, **kwargs):
        if cls._instance is None:
            cls._instance = super(ConfigManager, cls).__new__(cls)
        return cls._instance

    def __init__(self, *args: Union[str, Path], **kwargs: Any) -> None:
        if not ConfigManager._config: 
             ConfigManager._config = {}
             
        for arg in args:
            self.load(arg)
            
        if 'APP' not in ConfigManager._config:
            ConfigManager._config['APP'] = {}
            
        for key, value in kwargs.items():
            ConfigManager._config['APP'][key] = value

    def load(self, filename: Union[str, Path]) -> None:
        config = configparser.ConfigParser()
        config.read(filename)
        # Flatten to dict of dicts for easier access, keyed by section
        ConfigManager._config.update({s: dict(config.items(s)) for s in config.sections()})

    def get(self, key: str, section: str = 'APP', default: Any = None) -> Any:
        if section in ConfigManager._config and key in ConfigManager._config[section]:
            return ConfigManager._config[section][key]
        return default

    def set(self, key: str, value: Any, section: str = 'APP') -> None:
        if section not in ConfigManager._config:
            ConfigManager._config[section] = {}
        ConfigManager._config[section][key] = value
        
    @property
    def max_periods(self) -> int:
        return int(self.get('MAX_PERIODS', default=DEFAULT_MAX_PERIODS))
    
    @property
    def separator(self) -> str:
        return self.get('separator', default=DEFAULT_SEPARATOR)


class TeacherManager:
    """Handles loading and retrieving teacher details."""
    
    def __init__(self, workbook: Workbook):
        self.workbook = workbook
        self.teacher_details: Dict[str, Dict[str, str]] = {}
        self.class_incharge_map: Dict[str, str] = {}
        
    def load(self, ws_name: str = 'TEACHERS') -> None:
        if ws_name not in self.workbook:
            logger.warning(f"Sheet '{ws_name}' not found. Teacher details might be missing.")
            return

        ws = self.workbook[ws_name]
        headers = self._get_headers(ws)
        
        try:
            shortname_idx = headers.index('SHORTNAME')
        except ValueError:
            logger.error("SHORTNAME column not found in TEACHERS sheet.")
            return

        incharge_idx = -1
        if 'INCHARGE' in headers:
            incharge_idx = headers.index('INCHARGE')

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[shortname_idx] is None:
                continue
                
            shortname = str(row[shortname_idx]).strip()
            if not shortname or shortname.startswith('#'):
                continue
                
            details = {}
            for idx, header in enumerate(headers):
                if idx < len(row):
                    val = row[idx]
                    details[header] = str(val).strip() if val is not None else ""
            
            self.teacher_details[shortname] = details
            
            # Map class in-charge
            if incharge_idx != -1 and incharge_idx < len(row):
                klass = row[incharge_idx]
                if klass:
                    self.class_incharge_map[str(klass).strip()] = shortname

    def _get_headers(self, ws: Worksheet) -> List[str]:
        headers = []
        for cell in ws[1]:
            if cell.value is None:
                break
            headers.append(str(cell.value).strip())
        return headers

    def get_details(self, shortname: str) -> Dict[str, str]:
        return self.teacher_details.get(shortname, {})

    def get_name(self, shortname: str, fullname: bool = False) -> str:
        if fullname and shortname in self.teacher_details:
            return f"{self.teacher_details[shortname].get('NAME', shortname)}, {shortname}"
        return shortname
        
    def get_incharge_label(self, class_name: str) -> str:
        teacher_code = self.class_incharge_map.get(class_name)
        if not teacher_code:
            return "Class In-charge:" + '_' * 25
            
        details = self.get_details(teacher_code)
        if details:
            gender = details.get('GENDER', '')
            title = 'Ms' if gender.lower() in ['f', 'female'] else 'Mr'
            name = details.get('NAME', teacher_code)
            return f"Class In-charge: {title} {name}"
        return f"Class In-charge: {teacher_code}"

    def get_sorted_teachers(self, timetable_teachers: Set[str]) -> List[str]:
        # Defined teachers first, then any others found in timetable
        defined = [t for t in self.teacher_details if t in timetable_teachers]
        others = [t for t in timetable_teachers if t not in defined]
        return defined + others


class styles:
    """Excel style definitions."""
    GRAY_FILL = PatternFill(start_color="c3c3c3", end_color="c3c3c3", fill_type="solid")
    CENTER_ALIGN = Alignment(horizontal='center', vertical='top', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='top')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='top')
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    HEADER_FONT = Font(size=25)
    SUBHEADER_FONT = Font(size=16)


class TimetableManager:
    """Core logic for timetable processing."""
    
    def __init__(self, workbook: Workbook, config: ConfigManager, teacher_mgr: TeacherManager):
        self.workbook = workbook
        self.config = config
        self.teacher_mgr = teacher_mgr
        self.timetable: Dict[str, List[TimetableEntry]] = {}
        self.period_assignment_counts: Dict[str, int] = {} # teacher -> count

    def get_formatted_time(self) -> str:
        return time.ctime()

    def load_from_classwise(self) -> Tuple[int, int]:
        """Reads CLASSWISE sheet and populates self.timetable."""
        if "CLASSWISE" not in self.workbook:
            raise ValueError("CLASSWISE sheet not found.")
            
        sheet = self.workbook["CLASSWISE"]
        warnings = 0
        row_count = 0
        
        # Regex for cell content: Subject (Days) Teacher
        # Example: Math (1-3) SK
        pattern = re.compile(r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$')

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=10), start=2):
            class_cell = row[0]
            if not class_cell.value:
                break
                
            class_name = str(class_cell.value).strip()
            if class_name.startswith("#"):
                continue

            row_count += 1
            logger.info(f"Class: {class_name}...")
            # Normalize class name in cell
            class_cell.value = class_name
            
            course_period_counts: Dict[str, int] = {} # subject -> total periods
            
            # Columns 2 to 9 correspond to Periods 1 to 8
            for col_idx in range(1, 9): 
                cell = row[col_idx]
                content = cell.value
                
                if not content:
                    warnings += 1
                    logger.warning(f"Warning: Cell {get_column_letter(col_idx + 1)}{row_idx} is empty.")
                    continue
                
                cell_warnings = self._process_class_cell(
                    str(content), row_idx, col_idx + 1, pattern, class_name, course_period_counts
                )
                warnings += cell_warnings

            self._write_row_summary(sheet, row_idx, course_period_counts)

        # Update timestamp
        if not self.config.get('keepstamp', default=False):
             sheet.cell(row=row_count + 3, column=2).value = "Last updated on " + self.get_formatted_time()

        return row_count, warnings

    def _process_class_cell(self, content: str, row: int, col: int, pattern: re.Pattern, 
                           class_name: str, course_counts: Dict[str, int]) -> int:
        warnings = 0
        days_assigned: List[int] = []
        separator = self.config.separator
        
        for line in content.split(separator):
            line = line.strip()
            if not line or line.startswith("#"):
                continue

            match = pattern.match(line.upper())
            if not match:
                logger.warning(f"Formatting issue in Cell {get_column_letter(col)}{row}: '{line}'")
                return 1

            subject, days_str, teacher = match.groups()
            subject = subject.strip()
            
            expanded = DayUtils.expand_days(days_str)
            days_assigned.extend(expanded)
            
            count = DayUtils.count_days(days_str)
            course_counts[subject] = course_counts.get(subject, 0) + count

            period = col - 1
            entry = TimetableEntry(period, class_name, days_str, subject, teacher)
            self.timetable.setdefault(teacher, []).append(entry)

        # Check for missing days
        days_set = set(days_assigned)
        expected_days = {1, 2, 3, 4, 5, 6}
        if days_set != expected_days:
             missing = list(expected_days - days_set)
             if len(missing) < 6:
                  logger.warning(f"Warning: Missing days {missing} in cell {get_column_letter(col)}{row}.")
                  warnings += 1
        return warnings

    def _write_row_summary(self, sheet: Worksheet, row: int, counts: Dict[str, int]) -> None:
        summary = [f"{subj}: {count}" for subj, count in sorted(counts.items())]
        total = sum(counts.values())
        summary.append(f"TOTAL: {total}")
        sheet.cell(row=row, column=10).value = ", ".join(summary)

    def generate_teacherwise(self) -> None:
        """Generates the TEACHERWISE sheet."""
        if "TEACHERWISE" in self.workbook:
            ws = self.workbook["TEACHERWISE"]
            # Clear existing data
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row)
        else:
            logger.info("Creating TEACHERWISE sheet...")
            ws = self.workbook.create_sheet(title="TEACHERWISE", index=1)

        # Headers
        header = ["Name", 1, 2, 3, 4, 5, 6, 7, 8, "Periods", "Periods Daywise"]
        for col, val in enumerate(header, start=1):
            ws.cell(row=1, column=col).value = f"Period {val}" if isinstance(val, int) else val

        teachers = self.teacher_mgr.get_sorted_teachers(set(self.timetable.keys()))
        fullname_opt = self.config.get('fullname', default=False)
        separator = self.config.separator

        row = 2
        for teacher in teachers:
            # Col 1: Name
            ws.cell(row, 1).value = self.teacher_mgr.get_name(teacher, fullname_opt)
            
            entries = self.timetable.get(teacher, [])
            entries.sort(key=lambda x: (x.period, x.days)) # Sort by period then days
            
            period_cells: Dict[int, List[str]] = {}
            
            # Aggregate entries
            # We need to count periods accurately (handling multi-day entries)
            # Logic: sum of days for all entries
            
            period_day_map: Dict[int, Set[int]] = {} # period -> set of days
            day_period_map: Dict[int, Set[int]] = {} # day -> set of periods (for daywise count)

            for entry in entries:
                # Content for cell
                entry_str = f"{entry.class_name.strip()} ({entry.days}) {entry.subject}"
                period_cells.setdefault(entry.period, []).append(entry_str)
                
                # Counting logic
                expanded = entry.expanded_days
                period_day_map.setdefault(entry.period, set()).update(expanded)
                for d in expanded:
                    day_period_map.setdefault(d, set()).add(entry.period)
            
            # Fill cells
            for p_idx in range(1, 9):
                if p_idx in period_cells:
                    ws.cell(row, p_idx + 1).value = separator.join(period_cells[p_idx])
            
            # Total periods count
            total_count = sum(len(days) for days in period_day_map.values())
            ws.cell(row, 10).value = total_count
            
            # Daywise count
            day_counts = {d: len(day_period_map.get(d, set())) for d in range(1, 7)}
            ws.cell(row, 11).value = repr(day_counts)[1:-1] # strip {}

            row += 1

        if not self.config.get('keepstamp', default=False):
            ws.cell(row=len(teachers) + 3, column=2).value = "Last updated on " + self.get_formatted_time()

    def highlight_clashes(self) -> int:
        """Highlights clashes in TEACHERWISE sheet."""
        if "TEACHERWISE" not in self.workbook:
            return 0
        
        ws = self.workbook["TEACHERWISE"]
        separator = self.config.separator
        clash_mark = '**CLASH** '
        total_clashes = 0
        noclash = self.config.get('noclash', default=False)
        
        # Regex to parse back the cell content created in generate_teacherwise
        # Format: Class (Days) Subject
        p = re.compile(r'^(?P<class_name>[\w]+)\s*\((?P<days>.*)\)\s*(?P<subject>[\w \-.]+)$')

        for row in ws.iter_rows(min_row=2, max_col=9):
            teacher_name = row[0].value
            if not teacher_name: break
            
            for cell in row[1:]:
                if not cell.value: continue
                
                content_str = str(cell.value)
                lines = content_str.split(separator)
                day_class_map: Dict[int, List[str]] = {}
                
                for line in lines:
                    line = line.strip()
                    if not line or line.startswith('**CLASH**'): continue # Avoid re-parsing clash
                    
                    m = p.match(line)
                    if not m:
                        continue
                    
                    class_name, days_str, subject = m.groups()
                    try:
                        for day in DayUtils.expand_days(days_str):
                            day_class_map.setdefault(day, []).append(f"{class_name}-{subject}")
                    except ValueError:
                        continue
                        
                clash_days = [d for d, classes in day_class_map.items() if len(set(classes)) > 1]
                
                if clash_days:
                    total_clashes += len(clash_days)
                    clash_msg = repr(clash_days)
                    if not noclash:
                        # Prepend clash warning if not already there
                        if not str(cell.value).startswith(clash_mark):
                            cell.value = f"{clash_mark}{clash_msg}:\n{content_str}"
                    else:
                        logger.info(f"{clash_mark} {clash_msg} in {cell.coordinate}: {teacher_name}")
                        
        return total_clashes

    def generate_vacant_sheet(self) -> None:
        """Generates VACANT sheet."""
        if "TEACHERWISE" not in self.workbook: return
        src = self.workbook["TEACHERWISE"]
        
        if "VACANT" in self.workbook.sheetnames:
            ws = self.workbook["VACANT"]
            if ws.max_row >= 2: ws.delete_rows(2, ws.max_row)
        else:
            ws = self.workbook.create_sheet("VACANT")
            
        day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
        ws.cell(1, 1, "Teacher")
        for i in range(1, 7):
            ws.cell(1, i+1, day_names[i])
            
        row_idx = 2
        for row in src.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            if len(row) < 11 or not row[10]: continue
            
            teacher = row[0]
            daywise_str = str(row[10]) # e.g. "1: 4, 2: 5..."
            
            counts = {}
            for part in daywise_str.split(','):
                if ':' in part:
                    d, c = part.split(':')
                    try: counts[int(d)] = int(c)
                    except: pass
            
            ws.cell(row_idx, 1, teacher)
            max_p = self.config.max_periods
            for d in range(1, 7):
                ws.cell(row_idx, d+1, max_p - counts.get(d, 0))
            row_idx += 1

    def generate_adjustment_helper_sheet(self) -> None:
        """Generates the FREE_TEACHERS sheet."""
        if "FREE_TEACHERS" in self.workbook.sheetnames:
            ws = self.workbook["FREE_TEACHERS"]
            # Clear previous content (specifically the table area)
            for row in ws.iter_rows(min_row=2, max_row=8, min_col=2, max_col=self.config.max_periods + 1):
                for cell in row:
                    cell.value = None
        else:
            ws = self.workbook.create_sheet("FREE_TEACHERS")

        ws.cell(row=1, column=2, value='Free Teachers Sheet')
        FIRST_ROW = 2
        max_periods = self.config.max_periods

        # Write header: periods on top
        ws.cell(row=FIRST_ROW, column=1, value="Day/Period")
        for period in range(1, max_periods + 1):
            ws.cell(row=FIRST_ROW, column=period + 1, value=f"Period {period}")

        # Write days on left
        for day in range(1, 7):
            ws.cell(row=FIRST_ROW + day, column=1, value=f"Day {day}")

        # Build a lookup: teacher -> day -> set of busy periods
        teacher_busy_periods: Dict[str, Dict[int, Set[int]]] = {}
        for teacher, entries in self.timetable.items():
            teacher_busy_periods[teacher] = {}
            for entry in entries:
                for day in entry.expanded_days:
                    teacher_busy_periods[teacher].setdefault(day, set()).add(entry.period)

        # For each day and period, find free teachers
        for day in range(1, 7):
            for period in range(1, max_periods + 1):
                free_teachers = []
                for teacher in self.timetable:
                    busy_periods = teacher_busy_periods.get(teacher, {}).get(day, set())
                    if period not in busy_periods:
                        free_teachers.append(teacher)
                
                # Sort free_teachers by number of free periods (descending)
                # Note: teacher_busy_periods might not have an entry for a teacher on a specific day if they are fully free, 
                # but we initialized keys in the outer loop. Wait, we initialized only if they have entries.
                # Let's make it robust.
                
                def get_free_count(t, d):
                    busy = teacher_busy_periods.get(t, {}).get(d, set())
                    return max_periods - len(busy)

                free_teachers_sorted = sorted(
                    free_teachers,
                    key=lambda t: get_free_count(t, day),
                    reverse=True
                )
                
                # Format
                formatted = [
                    f"{t}:{get_free_count(t, day)}" for t in free_teachers_sorted
                ]
                ws.cell(row=day + FIRST_ROW, column=period + 1, value=", ".join(formatted))

        if not self.config.get('keepstamp', default=False):
            ws.cell(row=FIRST_ROW + 7, column=2).value = "Last updated on " + self.get_formatted_time()
        
        logger.info(f"Free teachers sheet written to 'FREE_TEACHERS'.")

    def convert_timetable_to_classwise_files(self, output_filename: str):
        """Generates individual class sheets in a new workbook."""
        output_path = Path(output_filename)
        if output_path.exists():
            if not self.config.get('yes', default=False):
                print(f"File {output_filename} already exists.")
                ans = input("Do you want to overwrite? [y/n]: ").strip().lower()
                if ans != 'y':
                    logger.info("Aborted by user.")
                    sys.exit(0)

        wb_out = Workbook()
        # Create Master template
        ws_master = wb_out.active
        ws_master.title = "MASTER"
        self._setup_classwise_template(ws_master)
        
        # We need to iterate the ORIGINAL CLASSWISE source again or rely on parsing
        # It's safer to read the source CLASSWISE sheet again to preserve order and formatting
        # or just map from our internal `timetable` if we grouped by class.
        # But `timetable` is grouped by teacher.
        # Let's read from source to be safe and simple.
        
        src_ws = self.workbook['CLASSWISE']
        row = 2
        
        # Regex for content
        p = re.compile(r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$')
        separator = self.config.separator
        
        while True:
            c_val = src_ws.cell(row, 1).value
            if not c_val: break
            
            class_name = str(c_val).strip()
            if class_name.startswith('#'):
                row += 1
                continue
                
            # Create sheet
            if class_name in wb_out: del wb_out[class_name]
            ws = wb_out.copy_worksheet(ws_master)
            ws.title = class_name
            
            # Header info
            ws.cell(2, 1).value = f"Class: {class_name}"
            ws.cell(2, 5).value = self.teacher_mgr.get_incharge_label(class_name)
            
            # Fill data
            for col in range(2, 10):
                content = src_ws.cell(row, col).value
                if not content: continue
                
                for line in str(content).split(separator):
                    line = line.strip()
                    if not line or line.startswith('#'): continue
                    
                    m = p.match(line)
                    if not m: continue
                    
                    subj, days_str, teacher = m.groups()
                    days = DayUtils.expand_days(days_str)
                    
                    for day in days:
                        # target row in template: day + 3 (Mon is row 4, day 1)
                        r_idx = day + 3 
                        cell = ws.cell(r_idx, col)
                        curr = cell.value if cell.value else ""
                        cell.value = f"{curr}{subj} ({teacher})\n"

            # Cleanup
            for r in range(4, 10):
                for c in range(2, 10):
                    cell = ws.cell(r, c)
                    if cell.value: cell.value = str(cell.value).strip()

            row += 1
            
        del wb_out['MASTER']
        wb_out.save(output_filename)

    def _setup_classwise_template(self, ws: Worksheet):
        ws['A4'], ws['A5'], ws['A6'] = 'Mon', 'Tue', 'Wed'
        ws['A7'], ws['A8'], ws['A9'] = 'Thu', 'Fri', 'Sat'
        for col in range(2, 10):
            ws.cell(3, col).value = col - 1
            
        ws.column_dimensions['A'].width = 16
        for col in range(2, 10):
            ws.column_dimensions[get_column_letter(col)].width = 14
        for r in range(1, 4): ws.row_dimensions[r].height = 34
        for r in range(4, 10): ws.row_dimensions[r].height = 54
        
        # Styles
        for col in range(1, 10):
            ws[f"{get_column_letter(col)}3"].fill = styles.GRAY_FILL
        for r in range(4, 10):
            ws[f"A{r}"].fill = styles.GRAY_FILL
            
        ws.merge_cells('A1:I1')
        ws.merge_cells('A2:D2')
        ws.merge_cells('E2:I2')
        
        ws['A1'].font = styles.HEADER_FONT
        ws['A1'].alignment = styles.CENTER_ALIGN
        ws['A1'].value = self.config.get('NAME', 'SCHOOL', 'Your School Name')
        
        ws['A2'].font = styles.SUBHEADER_FONT
        ws['E2'].font = styles.SUBHEADER_FONT
        ws['A2'].alignment = styles.LEFT_ALIGN
        ws['E2'].alignment = styles.RIGHT_ALIGN
        
        for r in range(3, 10):
            for c in range(1, 10):
                cell = ws.cell(r, c)
                cell.border = styles.THIN_BORDER
                cell.alignment = styles.CENTER_ALIGN


class DiffEngine:
    @staticmethod
    def compare(base_path: str, current_path: str) -> int:
        wb_base = openpyxl.load_workbook(base_path)
        wb_curr = openpyxl.load_workbook(current_path)
        
        if 'CLASSWISE' not in wb_base or 'CLASSWISE' not in wb_curr:
            logger.error("CLASSWISE sheet missing.")
            return 0
            
        ws_base = wb_base['CLASSWISE']
        ws_curr = wb_curr['CLASSWISE']
        
        diffs = []
        # Simple cell-by-cell comparison
        # Assuming structure is identical
        for row in range(2, ws_base.max_row + 1):
            class_val = ws_base.cell(row, 1).value
            if not class_val: break
            
            for col in range(2, 10):
                v1 = ws_base.cell(row, col).value
                v2 = ws_curr.cell(row, col).value
                if v1 != v2:
                    coord = f"{get_column_letter(col)}{row}"
                    diffs.append(coord)
                    ws_curr.cell(row, col).fill = styles.GRAY_FILL
        
        if diffs:
            logger.info(f"Differences found in: {', '.join(diffs)}")
            wb_curr.save(current_path)
        else:
            logger.info("No differences found.")
            
        return len(diffs)


class TwigApp:
    def __init__(self):
        self.config = ConfigManager()
        self.parser = self._create_parser()
        
    def _create_parser(self):
        parser = argparse.ArgumentParser(prog='twig.py', description='TeacherWise Timetable Generator')
        parser.add_argument('-i', '--config', default='twig.ini', help='config file')
        parser.add_argument('-k', '--keepstamp', action='store_true', help='keep time stamp')
        parser.add_argument('-s', '--separator', default="\n", help='separator')
        
        subparsers = parser.add_subparsers(dest="command", help="Subcommands")
        
        # Teacherwise
        tw = subparsers.add_parser("teacherwise")
        tw.add_argument('-f', '--fullname', action='store_true', help='use full names')
        tw.add_argument('-c', '--noclash', action='store_true', help='suppress clash marks')
        tw.add_argument('--dry-run', action='store_true', help='do not save files')
        tw.add_argument("infile", type=str, nargs='?', default='Timetable.xlsx')
        tw.add_argument("-o", "--outfile", type=str)
        
        # Classwise
        cw = subparsers.add_parser("classwise")
        cw.add_argument("-y", "--yes", action="store_true", help="overwrite without prompt")
        cw.add_argument("infile", type=str)
        cw.add_argument("outfile", type=str)

        # Diff
        df = subparsers.add_parser("diff")
        df.add_argument("base", type=str)
        df.add_argument("current", type=str)
        
        return parser

    def run(self):
        args = self.parser.parse_args()
        
        # Init Config
        self.config.load(args.config)
        self.config.set('separator', args.separator.replace('\\n', '\n'))
        self.config.set('keepstamp', args.keepstamp)
        
        if args.command == 'teacherwise':
            self._run_teacherwise(args)
        elif args.command == 'classwise':
            self.config.set('yes', args.yes)
            self._run_classwise(args)
        elif args.command == 'diff':
            DiffEngine.compare(args.base, args.current)
        else:
            self.parser.print_help()

    def _run_teacherwise(self, args):
        self.config.set('fullname', args.fullname)
        self.config.set('noclash', args.noclash)
        
        filename = args.infile
        logger.info(f"Loading {filename}...")
        try:
            wb = openpyxl.load_workbook(filename)
        except Exception as e:
            logger.error(f"Error loading {filename}: {e}")
            return

        teacher_mgr = TeacherManager(wb)
        teacher_mgr.load()
        
        tt_mgr = TimetableManager(wb, self.config, teacher_mgr)
        logger.info("Processing Classwise data...")
        tt_mgr.load_from_classwise()
        
        logger.info("Generating Teacherwise sheet...")
        tt_mgr.generate_teacherwise()
        
        logger.info("Checking clashes...")
        clashes = tt_mgr.highlight_clashes()
        logger.info(f"Clashes found: {clashes}")
        
        
        tt_mgr.generate_vacant_sheet()
        tt_mgr.generate_adjustment_helper_sheet()
        
        if args.dry_run:
            logger.info("Dry run: not saving.")
        else:
            outfile = args.outfile if args.outfile else filename
            logger.info(f"Saving to {outfile}...")
            wb.save(outfile)
            
    def _run_classwise(self, args):
        self.config.set('separator', args.separator.replace('\\n', '\n'))
        filename = args.infile
        wb = openpyxl.load_workbook(filename)
        teacher_mgr = TeacherManager(wb)
        teacher_mgr.load()
        
        tt_mgr = TimetableManager(wb, self.config, teacher_mgr)
        tt_mgr.convert_timetable_to_classwise_files(args.outfile)
        logger.info(f"Classwise sheets saved to {args.outfile}")


def main():
    app = TwigApp()
    try:
        app.run()
    except KeyboardInterrupt:
        sys.exit(1)

if __name__ == "__main__":
    main()
