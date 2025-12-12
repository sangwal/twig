#!/usr/bin/env python3

"""
    twig.py -- TeacherWIse [timetable] Generator

    A python script to generate Teacherwise timetable from Classwise
    timetable and individual classwise sheets for all classes.
"""

import argparse
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Set, Tuple, Optional
import configparser

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

__version__ = '251208'
MAX_PERIODS = 8
MAX_DAYS = 6


# ============================================================================
# Data Classes
# ============================================================================

@dataclass
class Args:
    """Command-line arguments container"""
    command: str
    infile: str
    outfile: Optional[str] = None
    config: str = 'twig.ini'
    keepstamp: bool = False
    separator: str = '\n'
    version: bool = False
    verbose: bool = False
    dry_run: bool = False
    fullname: bool = False
    noclash: bool = False
    yes: Optional[str] = None


@dataclass
class TeacherDetails:
    """Teacher information"""
    shortname: str
    name: str
    gender: str
    post: str
    incharge: Optional[str] = None
    other_fields: Dict = None

    def __post_init__(self):
        if self.other_fields is None:
            self.other_fields = {}


@dataclass
class TimetableEntry:
    """Single timetable entry"""
    period: int
    class_name: str
    days: str
    subject: str


@dataclass
class Context:
    """Application context"""
    args: Args
    config: 'Config'
    book: Optional[openpyxl.Workbook] = None


# ============================================================================
# Configuration
# ============================================================================

class Config:
    """Configuration manager"""
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super().__new__(cls)
            cls._instance._data = {}
        return cls._instance

    def load(self, filename: str) -> None:
        """Load configuration from file"""
        config = configparser.ConfigParser()
        config.read(filename)
        self._data = {section: dict(config[section]) for section in config.sections()}

    def get(self, key: str, section: str = 'APP', default=None):
        """Get configuration value"""
        return self._data.get(section, {}).get(key, default)

    def set(self, key: str, value, section: str = 'APP') -> None:
        """Set configuration value"""
        if section not in self._data:
            self._data[section] = {}
        self._data[section][key] = value

    def __repr__(self) -> str:
        return str(self._data)


# ============================================================================
# Utility Functions
# ============================================================================

def expand_days(days: str) -> List[int]:
    """
    Expand day ranges into individual days.
    Example: "1-3, 5-6" → [1, 2, 3, 5, 6]
    """
    result = []
    for group in re.split(r',\s*', days):
        if '-' in group:
            start, end = map(int, group.split('-'))
            if end < start:
                start, end = end, start
            result.extend(range(start, end + 1))
        else:
            try:
                result.append(int(group))
            except ValueError:
                continue
    return result


def compress_days(days: List[int]) -> str:
    """
    Compress individual days into ranges.
    Example: [1, 2, 3, 5, 6] → "1-3, 5-6"
    """
    if not days:
        return ""
    
    days = sorted(set(days))
    ranges = []
    start = prev = days[0]
    
    for day in days[1:]:
        if day == prev + 1:
            prev = day
        else:
            ranges.append(f"{start}-{prev}" if start != prev else str(start))
            start = prev = day
    
    ranges.append(f"{start}-{prev}" if start != prev else str(start))
    return ", ".join(ranges)


def count_days(days: str) -> int:
    """Count unique days in range string"""
    return len(set(expand_days(days)))


def get_formatted_time() -> str:
    """Get cached timestamp"""
    if not hasattr(get_formatted_time, "_cached_time"):
        get_formatted_time._cached_time = time.ctime()
    return get_formatted_time._cached_time


def escape_special_chars(char: str) -> str:
    """Escape special characters for display"""
    replacements = {
        '\n': '\\n', '\t': '\\t', '\r': '\\r',
        '\b': '\\b', '\f': '\\f', '\v': '\\v',
        '\\': '\\\\', '\'': '\\\'', '\"': '\\"'
    }
    return replacements.get(char, char)


def get_user_input(valid_chars: str, prompt: str, default: Optional[str] | None) -> str:
    """Get validated user input"""
    
    if default is str:
        print(prompt, end='', flush=True)
        print(default)
        if len(default) > 1:
            return default[0]
        return default
    
    while True:
        print(prompt, end='', flush=True)
        response = input().strip()
        if len(response) == 0:
            continue
        if response in valid_chars:
            return response
        print("Invalid choice. Try again.")


def get_class_number(class_name: str) -> str:
    """Extract class number (remove section letter)"""
    return class_name[:-1] if class_name else class_name


# ============================================================================
# Teacher and Timetable Loading
# ============================================================================

def load_teacher_details(workbook: openpyxl.Workbook, ws_name: str = 'TEACHERS') -> Dict[str, TeacherDetails]:
    """Load teacher details from TEACHERS sheet"""
    if ws_name not in workbook:
        raise ValueError(f"Sheet '{ws_name}' not found in workbook.")

    ws = workbook[ws_name]
    
    # Find headers
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        headers.append(str(val).strip().upper())

    # Validate required columns
    if 'SHORTNAME' not in headers:
        raise ValueError("SHORTNAME column not found in TEACHERS sheet.")

    shortname_idx = headers.index('SHORTNAME')
    teachers = {}

    row = 2
    while True:
        shortname = ws.cell(row=row, column=shortname_idx + 1).value
        if not shortname or str(shortname).strip() == '':
            break

        shortname = str(shortname).strip()
        if shortname.startswith('#'):
            row += 1
            continue

        # Build teacher record
        details_dict = {}
        for idx, header in enumerate(headers):
            val = ws.cell(row=row, column=idx + 1).value
            details_dict[header] = str(val).strip() if val else None

        teacher = TeacherDetails(
            shortname=shortname,
            name=details_dict.get('NAME', ''),
            gender=details_dict.get('GENDER', ''),
            post=details_dict.get('POST', ''),
            incharge=details_dict.get('INCHARGE'),
            other_fields={k: v for k, v in details_dict.items() 
                         if k not in ['SHORTNAME', 'NAME', 'GENDER', 'POST', 'INCHARGE']}
        )
        teachers[shortname] = teacher
        row += 1

    return teachers


# ============================================================================
# Timetable Generation
# ============================================================================

def load_timetable(input_sheet, separator: str) -> Tuple[int, Dict, Dict, int]:
    """
    Load timetable from CLASSWISE sheet.
    
    Returns:
        (num_classes, timetable, total_periods, warnings)
    """
    timetable: Dict[str, List[TimetableEntry]] = {}
    warnings = 0
    days_in_week = {1, 2, 3, 4, 5, 6}

    pattern = re.compile(
        r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$'
    )

    row = 2
    while True:
        class_name = input_sheet.cell(row, 1).value
        if not class_name:
            break

        class_name = class_name.strip()
        if class_name.startswith("#"):
            row += 1
            continue

        print(f"Class: {class_name}... ", end="")
        input_sheet.cell(row, 1).value = class_name

        periods_assigned = {}
        days_assigned_all = set()

        for column in range(2, 10):
            content = input_sheet.cell(row, column).value
            if not content:
                warnings += 1
                print(f"Warning: Cell {get_column_letter(column)}{row} is empty.")
                continue

            cell_warnings, days_assigned = _process_class_cell(
                content, row, column, separator, pattern, timetable, 
                class_name, periods_assigned, days_in_week
            )
            warnings += cell_warnings
            days_assigned_all.update(days_assigned)

        # Write period summary to column 10
        _write_period_summary(input_sheet, row, periods_assigned)
        print("done.")
        row += 1

    total_periods = {t: _count_periods(t, timetable) for t in timetable}
    num_classes = row - 2

    return num_classes, timetable, total_periods, warnings


def _process_class_cell(content: str, row: int, column: int, separator: str, pattern,
                       timetable: Dict, class_name: str, periods_assigned: Dict,
                       days_in_week: Set[int]) -> Tuple[int, Set[int]]:
    """Process a single CLASSWISE cell"""
    warnings = 0
    days_assigned = set()

    for line in content.split(separator):
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        match = pattern.match(line.upper())
        if not match:
            print(f"Warning: Cell {get_column_letter(column)}{row} has formatting issue.")
            print("    >>> ", line)
            return 1, days_assigned

        subject, days, teacher = match.groups()
        subject = subject.strip()
        expanded_days = expand_days(days)
        days_assigned.update(expanded_days)

        # Update period counts
        periods_assigned[subject] = periods_assigned.get(subject, 0) + count_days(days)

        # Update timetable
        entry = TimetableEntry(period=column - 1, class_name=class_name, days=days, subject=subject)
        timetable.setdefault(teacher, []).append(entry)

    # Warn about missing days
    if days_assigned and days_assigned != days_in_week:
        missing_days = sorted(days_in_week - days_assigned)
        print(f"Warning: Missing days {missing_days} in cell {get_column_letter(column)}{row}.")
        warnings += 1

    return warnings, days_assigned


def _write_period_summary(sheet, row: int, periods_assigned: Dict) -> None:
    """Write subject-period summary to column 10"""
    summary = [f"{subj}: {count}" for subj, count in sorted(periods_assigned.items())]
    total = sum(periods_assigned.values())
    summary.append(f"TOTAL: {total}")
    sheet.cell(row=row, column=10).value = ", ".join(summary)


def _count_periods(teacher: str, timetable: Dict) -> int:
    """Count total unique periods for a teacher"""
    period_count: Dict[int, Set[int]] = {}
    for entry in timetable[teacher]:
        days = set(expand_days(entry.days))
        period_count.setdefault(entry.period, set()).update(days)

    return sum(len(days) for days in period_count.values())


def _count_periods_daywise(teacher: str, timetable: Dict) -> Dict[int, int]:
    """Count periods per day for a teacher"""
    day_periods: Dict[int, Set[int]] = {}
    for entry in timetable[teacher]:
        for day in expand_days(entry.days):
            day_periods.setdefault(day, set()).add(entry.period)

    return {day: len(day_periods.get(day, set())) for day in range(1, 7)}


# ============================================================================
# Teacherwise Sheet Generation
# ============================================================================

def generate_teacherwise(workbook: openpyxl.Workbook, context: Context) -> Tuple[Dict, int, Dict]:
    """Generate teacherwise timetable from CLASSWISE"""
    if "CLASSWISE" not in workbook:
        raise ValueError("CLASSWISE sheet not found.")

    input_sheet = workbook["CLASSWISE"]
    print("Reading teacher details from 'TEACHERS' sheet... ", end="")
    teacher_details = load_teacher_details(workbook)
    print("done.")

    num_classes, timetable, total_periods, warnings = load_timetable(
        input_sheet, context.args.separator
    )

    _write_teacherwise_sheet(workbook, timetable, teacher_details, total_periods, context)

    if not context.args.keepstamp:
        input_sheet.cell(row=num_classes + 2, column=2).value = f"Last updated on {get_formatted_time()}"

    return timetable, warnings, total_periods


def _write_teacherwise_sheet(workbook: openpyxl.Workbook, timetable: Dict,
                            teacher_details: Dict, total_periods: Dict, context: Context) -> None:
    """Write teacherwise sheet to workbook"""
    if "TEACHERWISE" in workbook:
        output_sheet = workbook["TEACHERWISE"]
    else:
        print("Creating TEACHERWISE sheet... ", end="")
        output_sheet = workbook.create_sheet(title="TEACHERWISE", index=1)
        print("done.")

    _clear_sheet(output_sheet)

    # Write header
    header = ["Name", 1, 2, 3, 4, 5, 6, 7, 8, "Periods", "Periods Daywise"]
    for col, val in enumerate(header, start=1):
        output_sheet.cell(row=1, column=col).value = f"Period {val}" if isinstance(val, int) else val

    # Sort teachers
    timetable_teachers = set(timetable.keys())
    sorted_teachers = [t for t in teacher_details if t in timetable_teachers]
    sorted_teachers.extend(t for t in timetable_teachers if t not in sorted_teachers)

    args = context.args
    separator = args.separator

    # Write each teacher
    row = 2
    for teacher_code in sorted_teachers:
        if args.fullname:
            name = teacher_details[teacher_code].name
            teacher_label = f"{name}, {teacher_code}"
        else:
            teacher_label = teacher_code

        output_sheet.cell(row, 1).value = teacher_label

        for entry in sorted(timetable[teacher_code], key=lambda x: x.days):
            col = entry.period + 1
            existing = output_sheet.cell(row, col).value
            line = f"{entry.class_name.strip()} ({entry.days}) {entry.subject}"
            output_sheet.cell(row, col).value = f"{existing}{separator}{line}" if existing else line

        output_sheet.cell(row, 10).value = total_periods[teacher_code]
        periods_daywise = _count_periods_daywise(teacher_code, timetable)
        output_sheet.cell(row, 11).value = repr(periods_daywise)[1:-1]

        row += 1

    if not args.keepstamp:
        output_sheet.cell(row=len(sorted_teachers) + 2, column=2).value = f"Last updated on {get_formatted_time()}"


def _clear_sheet(sheet) -> None:
    """Clear sheet content"""
    row = 2
    while sheet.cell(row=row, column=1).value:
        for column in range(1, 11):
            sheet.cell(row=row, column=column).value = ""
        row += 1


# ============================================================================
# Clash Detection
# ============================================================================

def highlight_clashes(sheet, context: Context) -> int:
    """Detect and mark clashes in TEACHERWISE sheet"""
    args = context.args
    separator = args.separator
    clash_mark = '**CLASH** '
    total_clashes = 0

    pattern = re.compile(r'^(?P<class_name>[\w]+)\s*\((?P<days>.*)\)\s*(?P<subject>[\w \-.]+)$')

    row = 2
    while True:
        teacher_name = sheet.cell(row=row, column=1).value
        if not teacher_name:
            break

        for column in range(2, 10):
            content = sheet.cell(row, column).value
            if not content:
                continue

            lines = content.split(separator)
            entry = {}

            for line in lines:
                line = line.strip()
                if not line:
                    continue

                m = pattern.match(line)
                if not m:
                    print(f"\nWarning: Cell {get_column_letter(column)}{row} has formatting issue.")
                    print("    >>> ", line)
                    continue

                class_name, days, subject = m.groups()
                subject = subject.strip()

                try:
                    days_list = expand_days(days)
                except Exception as e:
                    print(f"ERROR: Cell {get_column_letter(column)}{row}: {e}")
                    sys.exit(1)

                for day in days_list:
                    if day not in entry:
                        entry[day] = []
                    entry[day].append(f"{get_class_number(class_name)}-{subject}")

            # Check for clashes
            clash_days = [day for day, items in entry.items() if len(set(items)) > 1]

            if clash_days:
                total_clashes += len(clash_days)
                clash_days_str = repr(clash_days)

                if not args.noclash:
                    sheet.cell(row=row, column=column).value = \
                        f"{clash_mark}{clash_days_str}:\n{sheet.cell(row=row, column=column).value}"
                else:
                    print(f"{clash_mark}{clash_days_str} in cell {get_column_letter(column)}{row}:\n"
                          f"Teacher {teacher_name}: {sheet.cell(row=row, column=column).value}")

        row += 1

    return total_clashes


def find_teachers_with_multiple_periods_same_class_day(timetable: Dict) -> List[Tuple]:
    """Find teachers with >2 periods in same class on same day"""
    issues = []

    for teacher, entries in timetable.items():
        class_day_periods: Dict[Tuple, List] = {}

        for entry in entries:
            for day in expand_days(entry.days):
                key = (entry.class_name, day)
                if key not in class_day_periods:
                    class_day_periods[key] = []
                class_day_periods[key].append((entry.period, entry.subject))

        for (class_name, day), periods in class_day_periods.items():
            if len(periods) > 2:
                period_list = [f"P{p[0]} ({p[1]})" for p in periods]
                issues.append((teacher, class_name, day, len(periods), period_list))

    return issues


# ============================================================================
# Classwise Sheets Generation
# ============================================================================

def generate_classwise(input_book: openpyxl.Workbook, outfile: str, context: Context) -> int:
    """Generate individual classwise sheets"""
    args = context.args
    config = context.config
    outfile_path = Path(outfile)

    if outfile_path.exists():
        print(f"File {outfile} already exists.")
        if args.yes:
            args.yes = 'y'  # default to 'yes' if provided
        # else args.yes = False --- IGNORE ---
        response = get_user_input('ynYN', 'Do you want to overwrite? y)es   n)o: ', args.yes)
        if response.lower() == 'n':
            print('Stopping prematurely.')
            sys.exit(1)

    input_sheet = input_book['CLASSWISE']

    try:
        output_book = openpyxl.load_workbook(outfile)
    except Exception:
        output_book = openpyxl.Workbook()

    # Create/update MASTER sheet
    master_sheet = _setup_master_sheet(output_book, config)

    # Create class sheets
    row = 2
    while True:
        klass = input_sheet.cell(row, 1).value
        if not klass:
            break

        if klass in output_book:
            del output_book[klass]

        print(f"Creating sheet {klass}...")
        copy = output_book.copy_worksheet(master_sheet)
        copy.title = klass
        row += 1

    # Load teacher details
    teacher_details = load_teacher_details(input_book)
    class_incharge = _load_class_incharge(input_book)

    # Process timetable
    pattern = re.compile(r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$')
    warnings = _populate_classwise_sheets(input_book, output_book, input_sheet, 
                                         class_incharge, teacher_details, pattern, context)

    # Cleanup and save
    _cleanup_classwise_sheets(output_book, input_sheet)
    output_book.save(outfile)

    return warnings

def show_differences(base, current):
    """
        Shows difference between base and current timetables

        base    -- filename of the base timetable (.xlsx)
        current -- filename of the current timetable (.xlsx)
    """

    # load the two  workbooks
    wb_base = openpyxl.load_workbook(base)
    wb_current = openpyxl.load_workbook(current)

    ws_base = wb_base['CLASSWISE']
    ws_current = wb_current['CLASSWISE']

    differences = []
    affected_teachers = []
    row = 2
    while True:
        class_name = ws_base.cell(row, 1).value
        if class_name is None:
            break

        for col in range(1, 10):
            cell_name = f"{get_column_letter(col)}{row}"
            if ws_base.cell(row, col).value != ws_current.cell(row, col).value:
                differences.append(cell_name)
                # print(f"Difference in {cell_name}")
                teachers = get_affected_teachers(ws_base, ws_current, cell_name)
                # print(teachers)
                affected_teachers.extend(teachers)
                # color code the change in the current in ws_current
                ws_current.cell(row, col).fill = PatternFill(start_color="c3c3c3", end_color="c3c3c3", fill_type="solid")

        row += 1

    affected_teachers = set(affected_teachers)      # remove duplicates
    affected_teachers = list(affected_teachers)     # re-convert to list
    print("Differences found in cells: ", ', '.join(differences))
    print(f"Likely affected teachers are: {', '.join(affected_teachers)}.")

    # save the changes to "current" file
    wb_current.save(current)
    # return number of differences found
    return len(differences)


def _setup_master_sheet(output_book: openpyxl.Workbook, config: Config):
    """Create or get MASTER sheet"""
    if 'MASTER' not in output_book:
        master_sheet = output_book.create_sheet('MASTER')

        # Write structure
        for i, day in enumerate(['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'], start=4):
            master_sheet[f'A{i}'] = day

        for col in range(2, 10):
            master_sheet.cell(3, col).value = col - 1

        _format_master_sheet(master_sheet)
    else:
        master_sheet = output_book['MASTER']

    master_sheet['A1'] = config.get('NAME', 'SCHOOL')
    return master_sheet


def _load_class_incharge(workbook: openpyxl.Workbook) -> Dict[str, str]:
    """Load class incharge mapping"""
    teachers_sheet = workbook['TEACHERS']

    # build header->col mapping using the correct worksheet variable
    headers = {}
    for col in range(1, teachers_sheet.max_column + 1):
        val = teachers_sheet.cell(1, col).value
        if val:
            headers[val.strip().upper()] = col

    if 'SHORTNAME' not in headers:
        raise ValueError("SHORTNAME column not found in TEACHERS sheet.")

    incharge = {}
    row = 2
    while True:
        teacher_code = teachers_sheet.cell(row, headers['SHORTNAME']).value
        if not teacher_code:
            break

        klass = None
        if 'INCHARGE' in headers:
            klass = teachers_sheet.cell(row, headers['INCHARGE']).value

        if klass:
            incharge[str(klass).strip()] = str(teacher_code).strip()
        row += 1

    return incharge


def _populate_classwise_sheets(input_book, output_book, input_sheet, class_incharge,
                              teacher_details, pattern, context) -> int:
    """Populate individual classwise sheets"""
    warnings = 0
    args = context.args

    row = 2
    while True:
        class_name = input_sheet.cell(row, 1).value
        if not class_name:
            break

        sheet_name = class_name.strip()
        ws = output_book[sheet_name]

        # Set class info
        ws.cell(2, 1).value = f"Class: {class_name}"
        if sheet_name in class_incharge:
            teacher = teacher_details[class_incharge[sheet_name]]
            title = 'Ms' if teacher.gender.lower() == 'f' else 'Mr'
            ws.cell(2, 5).value = f"Class In-charge: {title} {teacher.name}"
        else:
            ws.cell(2, 5).value = "Class In-charge:" + '_' * 25

        # Process periods
        for column in range(2, 10):
            content = input_sheet.cell(row, column).value
            if not content:
                warnings += 1
                print(f"Warning: Cell {get_column_letter(column)}{row} is empty.")
                continue

            for line in content.split(args.separator):
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                m = pattern.match(line)
                if not m:
                    print(f"Warning: Cell {get_column_letter(column)}{row} has formatting issue.")
                    print("    >>> ", line)
                    warnings += 1
                    continue

                subject, days, teacher = m.groups()
                for day in expand_days(days):
                    r = day + 3
                    cell = ws.cell(r, column)
                    cell.value = (cell.value or '') + f"{subject} ({teacher})\n"

        row += 1

    return warnings


def _cleanup_classwise_sheets(output_book, input_sheet) -> None:
    """Remove trailing newlines and add timestamps"""
    for ws in output_book:
        if ws.title and ws.title[0].isdigit():
            for r in range(4, 10):
                for c in range(2, 10):
                    cell = ws.cell(r, c)
                    if cell.value:
                        cell.value = cell.value.rstrip('\n')

    # Add timestamp
    timestamp = input_sheet.cell(input_sheet.max_row, 2).value
    for ws in output_book:
        if ws.title and ws.title[0].isdigit():  # a class sheet names start with a digit
            ws.cell(10, 2).value = timestamp


def _format_master_sheet(ws) -> None:
    """Format MASTER sheet"""
    ws.column_dimensions['A'].width = 16
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14

    for row in range(1, 4):
        ws.row_dimensions[row].height = 34
    for row in range(4, 10):
        ws.row_dimensions[row].height = 54

    # Shading
    for col in range(1, 10):
        ws[f'{get_column_letter(col)}3'].fill = PatternFill(start_color="c3c3c3", fill_type="solid")
    for row in range(4, 10):
        ws[f'A{row}'].fill = PatternFill(start_color="c3c3c3", fill_type="solid")

    # Merge and formatting
    ws.merge_cells('A1:I1')
    ws.merge_cells('A2:D2')
    ws.merge_cells('E2:I2')

    ws['A1'].font = Font(size=25)
    ws['A2'].font = Font(size=16)
    ws['E2'].font = Font(size=16)

    alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws['A1'].alignment = alignment
    ws['A2'].alignment = Alignment(horizontal='left', vertical='top')
    ws['E2'].alignment = Alignment(horizontal='right', vertical='top')

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    for row in range(3, 10):
        for col in range(1, 10):
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = alignment


# ============================================================================
# Vacant Periods Sheet
# ============================================================================

def generate_vacant_sheet(book: openpyxl.Workbook, context: Context) -> bool:
    """Generate VACANT sheet with free periods"""
    vacant_sheet_name = "VACANT"

    if "TEACHERWISE" not in book:
        raise ValueError('TEACHERWISE sheet not found.')

    input_sheet = book["TEACHERWISE"]

    if vacant_sheet_name in book.sheetnames:
        ws = book[vacant_sheet_name]
    else:
        ws = book.create_sheet(vacant_sheet_name)

    day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

    for row_idx, row in enumerate(input_sheet.iter_rows(min_row=1, values_only=True), start=1):
        if row_idx == 1:
            ws.cell(row=1, column=1, value="Teacher")
            for col in range(2, 8):
                ws.cell(row=1, column=col, value=day_names[col - 1])
            continue

        try:
            data_str = row[10]
        except (IndexError, TypeError):
            continue

        if not data_str:
            continue

        ws.cell(row=row_idx, column=1, value=input_sheet.cell(row=row_idx, column=1).value)

        try:
            data = {}
            for item in data_str.split(","):
                col, val = item.split(":")
                data[int(col.strip())] = int(val.strip())

            for col, val in data.items():
                ws.cell(row=row_idx, column=col + 1, value=MAX_PERIODS - val)
        except ValueError:
            continue

    if not context.args.keepstamp:
        ws.cell(row=input_sheet.max_row + 2, column=2).value = f"Last updated on {get_formatted_time()}"

    return True


# ============================================================================
# Configuration File
# ============================================================================

def write_sample_config(filename: str) -> None:
    """Write sample configuration file"""
    config_content = """
; Configuration file for twig.py

[SCHOOL]
SHORTNAME = AP
NAME = Your School (District, State)
ADDRESS = Your School Address Line 1
CITY = Your City
STATE = Your State
PIN = 000000

[APP]
MAX_PERIODS = 8
MAX_DAYS = 6
VERBOSE = true
DEBUG = false
"""
    with open(filename, 'w') as f:
        f.write(config_content)
    print(f"Sample configuration file '{filename}' created.")


# ============================================================================
# Main Entry Point
# ============================================================================

def setup_argument_parser() -> argparse.ArgumentParser:
    """Setup command-line argument parser"""
    parser = argparse.ArgumentParser(
        prog='twig.py',
        description='Generate teacherwise timetable from classwise timetable.'
    )

    parser.add_argument('-i', '--config', default='twig.ini', help='Configuration file')
    parser.add_argument('-k', '--keepstamp', action='store_true', help='Keep timestamp intact')
    parser.add_argument('-s', '--separator', default='\n', help='Line separator (default: \\n)')
    parser.add_argument('-v', '--version', action='store_true', help='Show version')
    parser.add_argument('-b', '--verbose', action='store_true', help='Verbose output')
    parser.add_argument('-d', '--dry-run', action='store_true', help='Dry run mode')

    subparsers = parser.add_subparsers(dest='command', help='Subcommands')

    # teacherwise subcommand
    tw = subparsers.add_parser('teacherwise', help='Generate teacherwise timetable')
    tw.add_argument('-f', '--fullname', action='store_true', help='Use full names')
    tw.add_argument('-c', '--noclash', action='store_true', help='Suppress CLASH marks')
    tw.add_argument('infile', help='Input file')
    tw.add_argument('-o', '--outfile', help='Output file (default: overwrite infile)')

    # classwise subcommand
    cw = subparsers.add_parser('classwise', help='Generate classwise sheets')
    cw.add_argument('-y', '--yes', action='store_true', help='Overwrite without prompting')
    cw.add_argument('infile', help='Input file')
    cw.add_argument('outfile', help='Output file')

    # diff subcommand
    diff = subparsers.add_parser('diff', help='Compare timetables')
    diff.add_argument('base', help='Base file')
    diff.add_argument('current', help='Current file')

    return parser


def main() -> int:
    """Main entry point"""
    parser = setup_argument_parser()
    args = parser.parse_args()

    if args.version:
        print(f"twig.py version {__version__} by Sunil Sangwal")
        sys.exit(0)

    # Handle separator
    if args.separator.startswith("\\"):
        args.separator = {'n': '\n', 't': '\t'}.get(args.separator[1], ';')

    # Load config
    config_path = Path(args.config)
    if not config_path.exists():
        write_sample_config(args.config)
        print(f"Please edit '{args.config}' and run again.")
        return 1

    print(f"Using configuration from {args.config}...")
    config = Config()
    config.load(args.config)

    # Create context
    context = Context(args=args, config=config)

    start_time = time.time()

    try:
        if args.command in ['teacherwise', 'classwise', 'diff']:
            book = openpyxl.load_workbook(args.infile)
            context.book = book

            if args.command == 'teacherwise':
                timetable, warnings, _ = generate_teacherwise(book, context)
                highlight_clashes(book['TEACHERWISE'], context)

                issues = find_teachers_with_multiple_periods_same_class_day(timetable)
                if issues:
                    print("\n⚠️  Teachers with >2 periods in same class on same day:")
                    for teacher, class_name, day, count, periods in issues:
                        print(f"  • {teacher} → {class_name} on Day {day} ({count} periods):")
                        for p in periods:
                            print(f"    - {p}")

                generate_vacant_sheet(book, context)

                if not args.dry_run:
                    outfile = args.outfile or args.infile
                    book.save(outfile)
                    print(f"Teacherwise timetable saved to '{outfile}'.")

            elif args.command == 'classwise':
                warnings = generate_classwise(book, args.outfile, context)
                print(f"Classwise timetables saved to '{args.outfile}'.")

            if warnings:
                print(f"Warnings: {warnings}")

        elif args.command == 'diff':
            print(f"Comparing '{args.base}' with '{args.current}'...")
            differences = show_differences(args.base, args.current)
            print(f"Differences found: {differences}")

        else:
            parser.print_help()
            return 0

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1

    finally:
        elapsed = time.time() - start_time
        print(f"Finished in {elapsed:.3f} seconds.")

    return 0


if __name__ == '__main__':
    sys.exit(main())