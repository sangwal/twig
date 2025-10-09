#!/usr/bin/env python3

"""
    twig.py -- TeacherWIse [timetable] Generator
    
    A python script to generate Teacherwise timetable from Classwise timetable
    and individual classwise sheets for all classes.
    
    The classwise timetable is read from CLASSWISE sheet and the generated teacherwise
    timetable is saved in TEACHERWISE sheet of the same input workbook.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    Written by Sunil Sangwal (sunil.sangwal@gmail.com)
    Date written: 20-Apr-2022
    Last Modified: 06-Nov-2024
"""
import argparse
import re
import time
import configparser # now settings are in twig.ini
import openpyxl
from pathlib import Path

from openpyxl.styles import Alignment, Border, Side

# change styles
# alignment = Alignment(horizontal='general',
#     vertical='top',
#     text_rotation=0,
#     wrap_text=True,
#     shrink_to_fit=False,
#     indent=0)

from openpyxl.styles import Font
# Make the cell bold using
# bold_font = Font(bold=True)
# ws['A1'].font = bold_font

# to revert to normal font (without bold), use
# normal_font = Font(bold=False)
# ws['A1'].font = normal_font

# Change the font color to red
# red_font = Font(color="FF0000")
# ws['A1'].font = red_font

from openpyxl.styles import PatternFill
# Change the background color of the cell to yellow
# fill_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
# ws['A1'].fill = fill_color

from openpyxl.utils import get_column_letter


__version__ = '251010'    # twig.py version YYMMDD

# configuration variables before running the script

expand_names = False    # set this to True to write full names of teachers
MAX_PERIODS = 8       # maximum number of periods in a day

class Config:
    _config = {}

    def __init__(self, *args, **kwargs):
        if len(Config._config) > 0:
            # already initialized
            return
        Config._config = {}
        # print(args, kwargs)
        
        for arg in args:
            self._load(arg)
        # override with kwargs
        for key in kwargs:
            self._config['APP'][key] = kwargs[key]

    def _load(self, filename):
        config = configparser.ConfigParser()
        config.read(filename)
        Config._config = dict(config)  # convert to normal dictionary
        # print(Config._config)
        
        return None # self._config

    def get(self, key: str, section='APP', default=None):
        # print(Config._config)
        if key in Config._config[section]:
            return Config._config[section][key]
        else:
            return default
    
    def set(self, key: str, value, section='APP'):
        Config._config[section][key] = value

    def __repr__(self):
        return str(Config._config)


def escape_special_chars(c):
    replacements = {
        '\n': '\\n',
        '\t': '\\t',
        '\r': '\\r',
        '\b': '\\b',
        '\f': '\\f',
        '\v': '\\v',
        '\\': '\\\\',
        '\'': '\\\'',
        '\"': '\\"'
    }
    if c in replacements:
        return replacements[c]
    return c

# chatgpt version
def expand_days(days):
    """
        Parameter
            days : eg. "1-2, 3, 4-6"
        
        Returns:
            [1, 2, 3, 4, 5, 6]
    """
    ret = []
    for group in re.split(r',\s*', days):
        if '-' in group:
            start_day, end_day = map(int, group.split('-'))
            if end_day < start_day:
                start_day, end_day = end_day, start_day
            ret.extend(range(start_day, end_day + 1))
        else:
            try:
                ret.append(int(group))
            except ValueError:
                continue
    return ret

# chatgpt version
def compress_days(days):
    """
        Parameter:
            days -- a list containing days in expanded form eg [1,2,3,5,6]
        
        Returns:
            a string of the form "1-3, 5-6"
    """
    if not days:
        return ""
    days = sorted(set(days))
    ranges = []
    start = prev = days[0]
    for day in days[1:]:
        if day == prev + 1:
            prev = day
        else:
            ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
            start = prev = day
    ranges.append(f"{start}-{prev}" if start != prev else f"{start}")
    return ", ".join(ranges)

def count_days(days):
    return len(set(expand_days(days)))

def count_periods(teacher, timetable):
    period_count = {}
    for period_info in timetable[teacher]:
        column, class_name, days, subject = period_info
        # print(period_info)
        days = expand_days(days)
        if column not in period_count:
            period_count[column] = []

        period_count[column].extend(days)   # combine two lists

    total_periods = 0
    for column in period_count:
        periods = len(set(period_count[column]))   # remove duplicates
        total_periods += periods
    
    return total_periods

def count_periods_daywise(teacher, timetable):
    # Build mapping day -> set(periods) to avoid duplicate counting of same period on same day
    teacher_timetable = timetable.get(teacher, [])
    day_periods = {}
    for period, _class, days, subject in teacher_timetable:
        for day in expand_days(days):
            day_periods.setdefault(day, set()).add(period)

    # Return counts for days 1..6 (zero if no periods)
    periods_daywise = {day: len(day_periods.get(day, set())) for day in range(1, 7)}

    # print(f"Teacher {teacher}: periods daywise: {periods_daywise}")
    return periods_daywise

# def get_formatted_time():
#     # t = time.localtime()
#     # return f"{t.tm_year}{t.tm_mon:02d}{t.tm_mday:02d}{t.tm_hour:02d}{t.tm_min:02d}{t.tm_sec:02d}"
#     return "Last updated on " + time.ctime()
def get_formatted_time():
    """
        returns a cached time string
    """
    if not hasattr(get_formatted_time, "_cached_time"):
        get_formatted_time._cached_time = time.ctime() # strftime("%H:%M:%S")
    return get_formatted_time._cached_time

# alternate implementation without using pandas
def load_teacher_details(workbook, ws_name='TEACHERS'):
    """
        loads details of teachers from the TEACHERS sheet
        returns a dictionary of the form
            {teacher_code: {SHORTNAME: ..., NAME: ..., Post: ...}}
    """
    if ws_name not in workbook:
        raise Exception(f"Sheet '{ws_name}' not found in the workbook.")

    ws = workbook[ws_name]
    # Find header row (assume first row)
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        headers.append(str(val).strip())

    # Find index of SHORTNAME
    try:
        shortname_idx = headers.index('SHORTNAME')
    except ValueError:
        raise Exception("SHORTNAME column not found in TEACHERS sheet.")

    teacher_details = {}
    row = 2
    while True:
        shortname = ws.cell(row=row, column=shortname_idx + 1).value
        if shortname is None or str(shortname).strip() == '':
            break
        shortname = str(shortname).strip()
        if shortname.startswith('#'): # ignore entries starting with '#'
            row += 1
            continue
        details = {}
        for idx, header in enumerate(headers):
            val = ws.cell(row=row, column=idx + 1).value
            if val is not None:
                val = str(val).strip()
            details[header] = val
        teacher_details[shortname] = details
        row += 1

    return teacher_details

def get_class_number(_class):
    return _class[:len(_class) - 1] # remove section (for example, 'A' from '10A')

def highlight_clashes(sheet, context):
    """
        reads teacherwise timetable from the TEACHERWISE sheet and highlights possible clashes
        by prepending **CLASH** to the offending cell
    """
    args = context['ARGS']
    SEPARATOR = args.separator
    CLASH_MARK = '**CLASH** '
    total_clashes = 0

    # format of line is "CLASS (1-3,5-6) SUBJECT", e.g., 10A (1-2, 4) MATH
    p = re.compile(r'^(?P<class_name>[\w]+)\s*\((?P<days>.*)\)\s*(?P<subject>[\w \-.]+)$')
    
    row = 2
    while True:
        if not sheet.cell(row=row, column=1).value:
            break

        for column in range(2, 10):
            content = sheet.cell(row, column).value
            # skip empty cells in class timetable with a warning
            if not content:
                # cells in teacherwise timetable could be empty; just skip them
                # warnings += 1
                # print(f"Warning: Cell {get_column_letter(column)}{row} of teacherwise timetable is empty.")
                continue
            # content = content.replace('\n', ';')
            # lines = content.split(";")
            lines = content.split(SEPARATOR) # SEPARATOR is "\n" or ;
            
            entry = {}

            for line in lines:
                line = line.strip()
                if line == "":
                    # skip empty lines
                    continue

                m = p.match(line)
                if not m:
                    print(f"\nWarning: Cell {get_column_letter(column)}{row} in 'Teacherwise' timetable has formatting issue.")
                    print("    >>> ", line)
                    # warnings += 1
                    continue
                class_name, days, subject = m.groups()
                subject = subject.strip()
                try:
                    days = expand_days(days)
                except:
                    print(f"\nERROR: (row={row}, column={column}) (Cell {get_column_letter(column)}{row}) in 'Teacherwise' timetable has formatting issue")
                    print(f"content: {content}, line: {line}")
                    exit(1)
                
                """
                    Ex 1:
                        10A (1-2) MATH
                        10B (2-3) MATH
                    
                    is not a clash; but

                    Ex 2:
                        10A (1-2) MATH
                        9B (2-3) MATH

                    is a clash.

                    2: [10, 9]
                    In Example 2 above, 2nd period: classes 10 and 9 simultaneously is a clash
                """

                for day in days:
                    if not day in entry:
                        entry[day] = []
                    entry[day].append(get_class_number(class_name)+ '-' + subject)    # Eg., '10-SCI' (from 10A (1-6) SCI)
                    # the above code now ensures that the case "7A (1) PE, 7B (1-4) MATH" is marked as a clash

            # after all lines in a cell have been processed
            clash_days = []
            for day in entry:
                entry[day] = set(entry[day])    # remove duplicates
                if len(entry[day]) > 1:
                    # possible clash
                    clash_days.append(day)

            # if there are clashes, write them
            if len(clash_days) > 0:
                total_clashes += len(clash_days)
                # converts list [1, 2, 5] into a string
                clash_days = repr(clash_days)
                sheet.cell(row=row, column=column).value = CLASH_MARK + f"{clash_days}:\n" + sheet.cell(row=row, column=column).value

        row += 1

    return total_clashes

def clear_sheet(sheet):
    # clear the sheet before starting writing...
    row = 2
    while True:

        for column in range(1, 11):
            sheet.cell(row=row, column=column).value = ""

        row += 1
        if not sheet.cell(row=row, column=1).value:
            # we have reacher EOF
            break

    return


def generate_teacherwise(workbook, context):
    """
    Generate teacherwise timetable from the CLASSWISE timetable.

    Args:
        workbook: openpyxl Workbook object
        context: dict containing configuration such as SEPARATOR and ARGS

    Returns:
        timetable (dict): The generated timetable data structure.
        warnings (int): Number of warnings generated while processing.
        total_periods: total periods assigned to teachers
    """
    args = context['ARGS']
    SEPARATOR = args.separator
    # print(args)

    if "CLASSWISE" not in workbook:
        raise Exception("CLASSWISE sheet not found. Stopping.")

    input_sheet = workbook["CLASSWISE"]

    # Load teacher names if available
    teacher_names = {}
    if "TEACHERS" in workbook:
        print("Reading teacher details from 'TEACHERS' sheet... ", end="")
        # teacher_names = load_teacher_names(workbook)
        teacher_details = load_teacher_details(workbook)
        print("done.")

    # Build timetable (core logic moved to helper)
    num_classes, timetable, total_periods, warnings = load_timetable(input_sheet, SEPARATOR)

    # Write teacherwise sheet
    write_teacherwise_sheet(workbook, timetable, teacher_details, total_periods, context)

    # Update timestamp in CLASSWISE
    if not args.keepstamp:
        input_sheet.cell(row=num_classes + 2, column=2).value = "Last updated on " + get_formatted_time()

    return timetable, warnings, total_periods
    # end of generate_teacherwise()


# ----------------------------------------------------------
# Helper Functions
# ----------------------------------------------------------

def load_timetable(input_sheet, SEPARATOR):
    """
    Build the timetable dictionary from the CLASSWISE sheet.

    Returns:
        timetable (dict): {teacher: [(period, class_name, days, subject), ...]}
        warnings (int): number of warnings
        total_periods (dict): total periods per teacher
    """
    timetable = {}
    warnings = 0
    days_in_week = {1, 2, 3, 4, 5, 6}

    pattern = re.compile(
        r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$'
    )

    # print("Processing timetable ...")

    row = 2
    while True:
        class_name = input_sheet.cell(row, 1).value
        if not class_name:
            break  # no more rows
        class_name = class_name.strip()
        
        if class_name.startswith("#"):
            row += 1
            continue  # skip commented rows

        print(f"Class: {class_name}... ", end="")
        
        # a space after the class name in the CLASSWISE sheet is ignored
        # space after class name gave me a great deal of headache: 
        # the incharge name was not being printed properly in the classwise sheets

        input_sheet.cell(row, 1).value = class_name  # trim spaces
        periods_assigned = {}

        for column in range(2, 10):  # periods 1-8
            content = input_sheet.cell(row, column).value
            if not content:
                warnings += 1
                print(f"Warning: Cell {get_column_letter(column)}{row} is empty.")
                continue

            warnings += process_class_cell(
                content, row, column, SEPARATOR, pattern, timetable, class_name, periods_assigned, days_in_week
            )

        # Write subject-period summary in column 10
        write_period_summary(input_sheet, row, periods_assigned)

        print("done.")
        row += 1

    # Compute total periods per teacher
    total_periods = {t: count_periods(t, timetable) for t in timetable}

    num_classes = row - 2  # number of classes processed

    return num_classes, timetable, total_periods, warnings
# end of load_timetable()


def process_class_cell(content, row, column, SEPARATOR, pattern, timetable, class_name, periods_assigned, days_in_week):
    """
    Process a single CLASSWISE cell (one period block for a class).
    Returns number of warnings.
    """
    warnings = 0
    days_assigned = []

    for line in content.split(SEPARATOR):
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        match = pattern.match(line.upper())
        if not match:
            print(f"Warning: Cell {get_column_letter(column)}{row} has formatting issue.")
            print("    >>> ", line)
            return 1  # one warning

        subject, days, teacher = match.groups()
        subject = subject.strip()

        expanded_days = expand_days(days)
        days_assigned.extend(expanded_days)

        # Track subject → count of periods
        periods_assigned[subject] = periods_assigned.get(subject, 0) + count_days(days)

        # Track teacherwise timetable
        period = column - 1
        timetable.setdefault(teacher, []).append((period, class_name, days, subject))

    # Warn if some days are missing
    if set(days_assigned) != days_in_week:
        missing_days = list(days_in_week - set(days_assigned))
        print(f"Warning: Missing days {missing_days} in cell {get_column_letter(column)}{row}.")
        warnings += 1

    return warnings


def write_period_summary(sheet, row, periods_assigned):
    """
    Write the summary of subject-period counts into column 10 of CLASSWISE.
    """
    summary = [f"{subj}: {count}" for subj, count in sorted(periods_assigned.items())]
    total = sum(periods_assigned.values())
    summary.append(f"TOTAL: {total}")
    sheet.cell(row=row, column=10).value = ", ".join(summary)

    # end of write_period_summary()


def write_teacherwise_sheet(workbook, timetable, teacher_details, total_periods, context):  # context has SEPARATOR, args
    """
    Create or update the TEACHERWISE sheet with the timetable data.
    """
    # Prepare sheet
    if "TEACHERWISE" in workbook:
        output_sheet = workbook["TEACHERWISE"]
    else:
        print("Creating TEACHERWISE sheet... ", end="")
        output_sheet = workbook.create_sheet(title="TEACHERWISE", index=1)
        print("done.")

    clear_sheet(output_sheet)

    # Header
    header = ["Name", 1, 2, 3, 4, 5, 6, 7, 8, "Periods", "Periods Daywise"]
    for col, val in enumerate(header, start=1):
        output_sheet.cell(row=1, column=col).value = val

    # Teachers ordering
    timetable_teachers = set(timetable.keys())
    sorted_teachers = [t for t in teacher_details if t in timetable_teachers]
    sorted_teachers.extend(t for t in timetable_teachers if t not in sorted_teachers)

    args = context['ARGS']
    expand_names = args.fullname
    SEPARATOR = args.separator

    # Write each teacher's timetable
    row = 2
    for teacher_code in sorted_teachers:
        if expand_names:
            teacher_label = f"{teacher_details[teacher_code]['NAME']}, {teacher_code}"  \
                if teacher_code in teacher_details else teacher_code
        else:
            teacher_label = teacher_code
        output_sheet.cell(row, 1).value = teacher_label

        for period, class_name, days, subject in sorted(timetable[teacher_code], key=lambda x: x[2]):
            col = period + 1
            existing = output_sheet.cell(row, col).value
            entry = f"{class_name.strip()} ({days}) {subject}"
            output_sheet.cell(row, col).value = f"{existing}{SEPARATOR}{entry}" if existing else entry

        output_sheet.cell(row, 10).value = total_periods[teacher_code]
        
        # write the periods per day in the K column of TEACHERWISE sheet
        periods_daywise = count_periods_daywise(teacher_code, timetable)
        periods_daywise = repr(periods_daywise)[1:-1]
        output_sheet.cell(row, 11).value = periods_daywise

        row += 1 # the last line of the loop

    
    # Timestamp
    if not args.keepstamp:
        output_sheet.cell(row=len(sorted_teachers) + 2, column=2).value = "Last updated on " + get_formatted_time()

    # end of write_teacherwise_sheet()

def get_user_input(valid_chars:str, prompt: str):
    while True:
        response = input(prompt)
        if response in valid_chars:
            return response
        print("Invalid choice. Try again.")


# print('Your choice: ', get_user_input('yYnNcC', 'y)es  n)o  c)ancel? '))
# exit(0)

def generate_classwise(input_book, outfile, context):
    """
        generate individual sheets for all classes to be printed for fixing in classrooms
    """

    # check if outfile already exists. If it exists, prompt user for confirmation
    outfile_path = Path(outfile)
    if outfile_path.exists():
        # outfile already exists
        print(f"File {outfile} already exists.")
        response = get_user_input('ynYN', 'Do you want to overwrite? y)es   n)o: ')
        if response.lower() == 'n':    
            print('Stopping prematurely. Re-run with other filename.')
            exit(1)

    config = Config()

    master_sheet = None

    input_sheet = input_book['CLASSWISE']

    try:
        output_book = openpyxl.load_workbook(outfile) # Workbook()
    except:
        # create an empty book if there is no workbook already
        output_book = openpyxl.Workbook()
        
    if 'MASTER' not in output_book:
        master_sheet = output_book.create_sheet('MASTER')

        # write the header
        master_sheet['A4'] = 'Mon'
        master_sheet['A5'] = 'Tue'
        master_sheet['A6'] = 'Wed'
        master_sheet['A7'] = 'Thu'
        master_sheet['A8'] = 'Fri'
        master_sheet['A9'] = 'Sat'

        for col in range(2, 10):
            master_sheet.cell(3, col).value = col - 1   # periods 1 - 8

        format_master_ws(master_sheet)
    else:
        master_sheet = output_book['MASTER']
    
    # update school name irrespective of whether the sheet was newly created or already existed
    master_sheet['A1'] = config.get('NAME', 'SCHOOL')   # 'GSSS AMARPURA (FAZILKA)'

    # read the names of incharges from the TEACHERS sheet
    teachers_sheet = input_book['TEACHERS']
    class_incharge = {}
    
    # some settings!!
    MAX_TEACHER_FIELDS = 12

    # find the column indexes for fields in the TEACHERS sheet
    column_index = {}
    for col in range(1, MAX_TEACHER_FIELDS):
        cell_value = teachers_sheet.cell(1, col).value
        if cell_value is None or cell_value == '':
            break
        cell_value = cell_value.strip()
        column_index[cell_value] = col

    # read the incharge details
    row = 2
    while True:
        teacher_code = teachers_sheet.cell(row, column_index['SHORTNAME']).value
        if teacher_code is None or teacher_code == '':
            break
        klass = teachers_sheet.cell(row, column_index['INCHARGE']).value
        
        teacher_code = teacher_code.strip()
        klass = klass.strip() if klass is not None else ''

        if klass is not None and klass != '':
            class_incharge[klass] = teacher_code

        row += 1

    # copy/create templates for each class
    row = 2
    while True:
        klass = input_sheet.cell(row, 1).value
        if klass is None or klass == '':
            break

        # the following code effectively clears the sheet before writing any data
        if klass in output_book:
            # delete old one
            del output_book[klass]

        # create new by copying from the master
        print(f"creating sheet {klass} ...")
        copy = output_book.copy_worksheet(master_sheet)
        copy.title = klass

        row += 1

    # set up loops and process
    p = re.compile(r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$')

    teacher_details = load_teacher_details(input_book)
    
    warnings = 0
    row = 2
    while True:
        # print(f"Input Sheet: {input_sheet.title} row={row}")
        class_name = input_sheet.cell(row, 1).value
        if not class_name:
            break       # we have reached the end of CLASSWISE sheet, so stop further processing
        
        sheet_name = class_name
        # write class name
        output_book[sheet_name].cell(2, 1).value = f"Class: {class_name}"

        # write incharge name    
        if class_name in class_incharge:
            title = 'Ms' if teacher_details[class_incharge[class_name]]['GENDER'] in ['f', 'F'] else 'Mr'
            output_book[sheet_name].cell(2, 5).value = \
                f"Class In-charge: {title} {teacher_details[class_incharge[class_name]]['NAME']}"
        else:
            output_book[sheet_name].cell(2, 5).value = "Class In-charge:" + '_' * 25    # leave space for writing name of the incharge

        for column in range(2, 10):
            content = input_sheet.cell(row, column).value
            # skip empty cells in class timetable with a warning
            if not content:
                warnings += 1
                print(f"Warning: Cell {get_column_letter(column)}{row} is empty.")
                continue

            lines = content.split(context['ARGS'].separator) # SEPARATOR is "\n" or ;
            
            for line in lines:
                line = line.strip()
                if line == '' or line.startswith('#'):  # ignore empty lines and the ones starting with '#' -- used as comment
                    continue

                m = p.match(line)
                if m is None:   # no match
                    # print(f"\nWarning: (row={row}, column={column}) (Cell {get_column_letter(column)}{row}) has some formatting issue")
                    print(f"Warning: Cell {get_column_letter(column)}{row} in CLASSWISE sheet has some formatting issue.")
                    print("    >>> ", line)
                    warnings += 1
                    continue

                subject, days, teacher = m.groups()
                subject = subject.strip()
                days = expand_days(days)

                # copy data to the respective classwise sheet
                for day in days:
                    r = day + 3     # variable "row" is already taken
                    # print(f"output_book[{sheet_name}].cell({r}, {column}).value += {subject} ({teacher})")
                    if output_book[sheet_name].cell(r, column).value is None:
                        output_book[sheet_name].cell(r, column).value = ''
                    output_book[sheet_name].cell(r, column).value += f"{subject} ({teacher})\n"
        
        row += 1
        # end of while True loop

    # get the time stamp from the CLASSWISE sheet
    timestamp = input_sheet.cell(row, 2).value
    for ws in output_book:
        if ws.title[0].isdigit():
            ws.cell(10, 2).value = timestamp

    # save everything to the file
    output_book.save(outfile)
    
    return warnings
    # end generate_classwise(filename)

def get_teachers_in_cell(ws, cell_name):
    config = Config()
    p = re.compile(r'^(?P<subject>[\w \-.]+)\s*\((?P<days>[1-6,\- ]+)\)\s*(?P<teacher>[A-Z]+)$')
    content = ws[cell_name].value
    if config.get('ARGS'):
        lines = content.split(config.get('ARGS').separator) # SEPARATOR is "\n" or ;
    else:
        lines = content.split('\n') # SEPARATOR is "\n"
    teachers = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith('#'):
            continue
        m = p.match(line)
        if not m:
            print(lines)
            raise Exception(f"Error: {cell_name} is not in correct format.")
        subject, days, teacher = m.groups()
        teachers.append(teacher)

    return teachers

def get_affected_teachers(ws_base, ws_current, cell_name):
    # simplest implementation is to consider every teacher in the corresponding cells as affected
    
    # read names of teachers in both sheets
    teachers = []
    # first, read from base sheet
    teachers.extend(get_teachers_in_cell(ws_base, cell_name))
    teachers.extend(get_teachers_in_cell(ws_current, cell_name))
    teachers = list(set(teachers))    # remove duplicates

    return teachers   # re-convert to list
    # read from the current sheet
    
def show_differences(base, current):
    """
        Shows difference between base and current timetables

        base    -- filename of the base timetable (.xlsx)
        current -- filename of the current timetable (.xlsx)
    """

    # load the two  workbooks
    wb_base = openpyxl.load_workbook(base)
    wb_current = openpyxl.load_workbook(current)

    ws_base = wb_base['CLASSWISE']
    ws_current = wb_current['CLASSWISE']

    differences = []
    affected_teachers = []
    row = 2
    while True:
        class_name = ws_base.cell(row, 1).value
        if class_name is None:
            break

        for col in range(1, 10):
            cell_name = f"{get_column_letter(col)}{row}"
            if ws_base.cell(row, col).value != ws_current.cell(row, col).value:
                differences.append(cell_name)
                # print(f"Difference in {cell_name}")
                teachers = get_affected_teachers(ws_base, ws_current, cell_name)
                # print(teachers)
                affected_teachers.extend(teachers)
                # color code the change in the current in ws_current
                ws_current.cell(row, col).fill = PatternFill(start_color="c3c3c3", end_color="c3c3c3", fill_type="solid")

        row += 1

    affected_teachers = set(affected_teachers)  # remove duplicates
    affected_teachers = list(affected_teachers) # re-convert to list
    print("Differences found in cells: ", ', '.join(differences))
    print(f"Likely affected teachers are: ", ', '.join(affected_teachers)+'.')

    # save the changes to "current" file
    wb_current.save(current)
    # return number of differences found
    return len(differences)

def format_master_ws(ws):
    ws.column_dimensions['A'].width = 16 # first column
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14 # all other columns

    # first three rows
    for row in range(1, 4):
        ws.row_dimensions[row].height = 34
    
    # rows 4 to 9
    for row in range(4, 10):
        ws.row_dimensions[row].height = 54
    
    # shade the row showing periods (3rd row)
    for col in range(1, 10):
        ws[get_column_letter(col)+'3'].fill = PatternFill(start_color="c3c3c3", end_color="c3c3c3", fill_type="solid")
    # shade the days in Column A
    for row in range(4, 10):
        ws['A'+str(row)].fill = PatternFill(start_color="c3c3c3", end_color="c3c3c3", fill_type="solid")

    # format header
    ws.merge_cells('A1:I1')
    ws.merge_cells('A2:D2')
    ws.merge_cells('E2:I2')

    ws['A1'].font = Font(size=25)
    ws['A2'].font = Font(size=16)
    ws['E2'].font = Font(size=16)

    alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    ws['A1'].alignment = alignment # school name
    ws['A2'].alignment = Alignment(horizontal='left', vertical='top')   # Class
    ws['E2'].alignment = Alignment(horizontal='right', vertical='top')  # Incharge

    # Define the border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in range(3, 10):
        for col in range(1, 10):
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = alignment

    return
    # end format_master_ws()

def generate_vacant_sheet(book, context):
    """
        generates the VACANT sheet with number of free periods for each teacher on each day
    """
    VACANT_SHEET = "VACANT"

    if "TEACHERWISE" not in book:
        raise Exception('TEACHERWISE sheet not found. Stopping.')
    
    input_sheet = book["TEACHERWISE"]
    
    # Load workbook and sheet
    ws = input_sheet    # book["TEACHERWISE"]
    wb = book           # openpyxl.load_workbook(filename)

    # Create (or get) output sheet
    if VACANT_SHEET in wb.sheetnames:
        out_ws = wb[VACANT_SHEET]
    else:
        out_ws = wb.create_sheet(VACANT_SHEET)

    day_names = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
    # Loop over rows in TEACHERWISE
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if row_idx == 1:
            # Header row
            out_ws.cell(row=row_idx, column=1, value="Teacher")
            for col in range(2, 8):
                # write headers
                out_ws.cell(row=row_idx, column=col, value=day_names[col - 1])  # Days 1-6
            continue
        # ------ TODO: problematic code starts here -------
        try:
            data_str = row[10]  # 11th column (0-based index = 10)
        except Exception as e:
            continue
            print(len(row), row)
            print(e)
            exit(0)

        # -----end---- #
        
        if not data_str:
            continue  # skip empty cells

        # copy teacher name in first column
        out_ws.cell(row=row_idx, column=1, value=input_sheet.cell(row=row_idx, column=1).value)  # Teacher name

        # Parse string into dictionary
        data = {}
        for item in data_str.split(","):
            try:
                # print("item: ", item)
                col, val = item.split(":")
                data[int(col.strip())] = int(val.strip())
            except Exception as e:
                print(f"generate_vacant_sheet(): Runtime error: {e}")
                continue

        # Write into VACANT sheet (same row number)
        for col, val in data.items():
            out_ws.cell(row=row_idx, column=col+1, value=MAX_PERIODS - val)

    return True # generate_vacant_sheet()

# FREE_TEACHERS sheet:
# Teachers who are free on a particular day and period
def generate_adjustment_helper_sheet(timetable, context):
    """
        generate a sheet to help in adjusting timetable
    """
    # Create or get the sheet
    book = context.get('book', None)
    if not book:
        raise Exception("Workbook not found in context.")

    FREE_SHEET = "FREE_TEACHERS"
    if FREE_SHEET in book.sheetnames:
        ws = book[FREE_SHEET]
        # Clear previous content
        for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=9):
            for cell in row:
                cell.value = None
    else:
        ws = book.create_sheet(FREE_SHEET)

    ws.cell(row=1, column=2, value='Free Teachers Sheet')
    FIRST_ROW = 2

    # Write header: periods on top
    ws.cell(row=FIRST_ROW, column=1, value="Day/Period")
    for period in range(1, MAX_PERIODS+1):
        ws.cell(row=FIRST_ROW, column=period+1, value=f"Period {period}")

    # Write days on left
    for day in range(1, 7):
        ws.cell(row=FIRST_ROW + day, column=1, value=f"Day {day}")

    # Build a lookup: teacher -> day -> set of busy periods
    teacher_busy_periods = {}
    # timetable['AS'] equals
    # 'AS': [(2, '11A', '3-6', 'PBI'), (3, '11A', '6', 'PBI'), (5, '11A', '2', 'ROT HIS'), (7, '11A', '5-6', 'PBI'), ...]
    for teacher in timetable:
        teacher_busy_periods[teacher] = {}
        for period_info in timetable[teacher]:
            period, _, days, _ = period_info
            for day in expand_days(days):
                teacher_busy_periods[teacher].setdefault(day, set()).add(period)

    # For each day and period, find free teachers
    for day in range(1, 7):
        for period in range(1, MAX_PERIODS+1):
            free_teachers = []
            for teacher in timetable:
                busy_periods = teacher_busy_periods[teacher].get(day, set())
                if period not in busy_periods:
                    free_teachers.append(teacher)
            # Sort free_teachers by number of free periods (descending)
            free_teachers_sorted = sorted(
                free_teachers,
                key=lambda t: MAX_PERIODS - len(teacher_busy_periods[t].get(day, set())),
                reverse=True
            )
            # Format as "teacher_code : number_of_free_periods"
            formatted = [
                f"{t}:{MAX_PERIODS - len(teacher_busy_periods[t].get(day, set()))}"
                for t in free_teachers_sorted
            ]
            ws.cell(row=day+FIRST_ROW, column=period+1, value=", ".join(formatted))

    # Timestamp
    if not context['ARGS'].keepstamp:
        ws.cell(row=FIRST_ROW + 7, column=2).value = "Last updated on " +  get_formatted_time()

    verbose(f"Free teachers sheet written to '{FREE_SHEET}'.")

    return # generate_adjustment_helper_sheet()

def verbose(msg, level=1):
    if level > 1:
        print(msg)
    return

def write_sample_config(filename):
    CONFIG_FILE = filename
    DEFAULT_CONFIG = """[SCHOOL]
SHORTNAME = AP
NAME = Your School (District, State)
ADDRESS = Your School Address Line 1
ADDRESS2 = Your School Address Line 2
CITY = Your City
STATE = Your State
PIN = 000000
PHONE = +91-0000-000000
EMAIL = yourschoolname@gmail.com
WEBSITE = www.yourschool.in
LOGO = school_logo.png
; MOTTO = 
MOTTO_EN = Lead me from darkness to light
AFFILIATION = XXXXXX
UDISE = 00000000000

[PRINCIPAL]
NAME = principal name
DEGREE = MA, B.Ed.
DESIGNATION = Principal

[GENERATED_BY]
NAME = Sunil Sangwal
EMAIL = sunil.sangwal@gmail.com
WEBSITE = https://sangwal.in
GITHUB = https://www.github.com/sangwal/twig.git

; settings regarding Timetable or twig.py
[APP]
MAX_PERIODS = 8
MAX_DAYS = 6
OUTPUT_FILE = timetable.xlsx
INPUT_FILE = timetable.xlsx
LOG_FILE = timetable.log
RANDOM_SEED = 42
VERBOSE = true
DEBUG = true

[SECTION]
A = Daisy
B = Lotus
C = Marigold
D = Tulip
E = Rose
F = Sunflower
G = Jasmine
H = Hibiscus
I = Orchid
J = Lily
K = Daffodil
L = Poppy
M = Iris
N = Dahlia
O = Carnation
P = Chrysanthemum
Q = Gardenia
R = Azalea
S = Begonia
T = Camellia
U = Freesia
V = Gladiolus
W = Hydrangea
X = Lilac
Y = Magnolia
Z = Peony
"""
    # print(f"Configuration file '{CONFIG_FILE}' not found. Creating a new one with default settings.")
    with open(CONFIG_FILE, 'w') as config:
        config.write("; Configuration file for twig.py\n")
        config.write("; You can modify the settings as needed.\n\n")
        config.write(DEFAULT_CONFIG)
        config.close()
        print(f"Default configuration written to '{CONFIG_FILE}'. You can modify it as needed and rerun the program.")

def main():
    ##########################################################
    #
    # process command line arguments
    #
    #
    parser = argparse.ArgumentParser(prog='twig.py', description='Generates teacherwise (or classwise) timetable from classwise (or teacherwise) timetable.')
    parser.version = '1.0'

    parser.add_argument('-i', '--config', action='store',
        help='configuration file; default is twig.ini', default='twig.ini')
    parser.add_argument('-k', '--keepstamp', action='store_true', help='keep time stamp intact')
    parser.add_argument('-s', '--separator', action='store', help='newline separator; default is \\n', default="\n")
    parser.add_argument('-v', '--version', action='store_true', help='display version information')
    parser.add_argument('-b', '--verbose', action='store_true', help='verbose output')

    # Create a subparsers object
    subparsers = parser.add_subparsers(dest="command", help="Subcommands")

    # Subcommand 'teacherwise'
    tw_parser = subparsers.add_parser("teacherwise", help="Generate teacherwise timetable")
    tw_parser.add_argument('-f', '--fullname', action='store_true', help='replace short names with full names')
    tw_parser.add_argument("infile", type=str, action="store", help="File containing classwise timetable")

    # Subcommand 'classwise'
    cw_parser = subparsers.add_parser("classwise", help="Generate classwise timetable")
    cw_parser.add_argument("infile", type=str, action="store", help="File containing classwise timetable")
    cw_parser.add_argument("outfile", type=str, action="store", help="File to write classwise timetable")

    # # subcommand 'vacant'
    # vacant_parser = subparsers.add_parser("vacant", help="show vacant periods for all teachers")
    # vacant_parser.add_argument("infile", type=str, action="store", help="File containing classwise timetable")
    
    # Subcommand 'diff'    
    diff_parser = subparsers.add_parser("diff", help="compare two timetables")
    diff_parser.add_argument("base", type=str, action="store", help="base classwise timetable to compare against")
    diff_parser.add_argument("current", type=str, action="store", help="current timetable to be compared against base timetable")

    # Parse the arguments
    args = parser.parse_args()

    # print(args)
    if args.version:
        print(f"twig.py: version {__version__} by Sunil Sangwal")
        exit(0)

    # expand_names = getattr(args, "fullname", False)    # True or False; default = False

    # if not args.separator:
    #     args.separator = "\n"    # multi-line separator
    # else:
    #     # args.SEPARATOR = args.separator
    #     if args.separator == '\\n':
    #         args.separator = '\n'
    verbose(f"Using Separator '{escape_special_chars(args.separator)}' ...")

    startTime = time.time()

    # load settings from twig.ini file
    CONFIG_FILE = args.config # 'twig.ini'

    # using pathlib (modern, recommended)
    config_path = Path(CONFIG_FILE)
    if not config_path.exists():
        # create a default config file
        print(f"Configuration file '{CONFIG_FILE}' not found. Creating a new one with default settings.")
        write_sample_config(CONFIG_FILE)
        print(f"Sample configuration file '{CONFIG_FILE}' created. Please edit it as needed and run again.")
        exit(1)

    print(f"Using configuration from {CONFIG_FILE}...")

    config = Config(CONFIG_FILE)

    
    DEBUG = config.get('DEBUG')
    if DEBUG.lower() == 'true' or DEBUG == '1':
        DEBUG = True
    else:
        DEBUG = False
    verbose(f"DEBUG is {DEBUG}")

    warnings = 0

    if DEBUG:
        filename = "timetable.xlsx" # input file
        args.fullname = True
        args.keepstamp = False
        # args.separator = '\n'

    context = {
        'ARGS' : args
    }
 
    args.command = args.command.lower() if args.command else None

    if args.command in ['teacherwise', 'classwise', 'vacant']:
        if not args.infile:
            filename = 'Timetable.xlsx'
        else:
            filename = args.infile

        verbose(f"Reading CLASSWISE timetable from '{filename}'... ", level=2)
        book = openpyxl.load_workbook(filename)
        book.filename = filename    # remember the filename
        # print("done.")

    if args.command == 'classwise':
        verbose(f"Generating classwise timetable in '{args.outfile}' ...")
        # read classwise timetable and generate classwise sheets
        warnings = generate_classwise(book, args.outfile, context)
        verbose(f"Classwise timetables saved to '{args.outfile}'.")
        if warnings:
            print(f"Warnings: {warnings}")
    elif args.command == 'teacherwise':
        # read classwise timetable and generate teacherwise timetable
        context['book'] = book

        # print("context is: ", context)
        timetable, warnings, total_periods = generate_teacherwise(book, context)
        teacherwise_sheet = book['TEACHERWISE']
        
        # Highlight possible clashes
        verbose("Highlighting clashes in 'TEACHERWISE' sheet ...",)
        total_clashes = highlight_clashes(teacherwise_sheet, context)

        # generate vacant periods sheet as well
        verbose("Generating vacant periods sheet 'VACANT' ...", level=2)
        generate_vacant_sheet(book, context)

        # generate adjustment helper sheet as well
        verbose("Generating adjustment helper sheet 'FREE_TEACHERS' ...", level=2)
        generate_adjustment_helper_sheet(timetable, context)    # use timetable generated above

        # save the teacherwise timetable
        verbose(f"Saving teacherwise timetable to '{filename}' ...")
        book.save(filename)
        verbose(f"Teacherwise timetable saved to 'TEACHERWISE' sheet of '{filename}'.", level=2)

        verbose(f"Clashes: {total_clashes}", level=2)
        verbose(f"Warnings: {warnings}", level=2)

    # elif args.command == 'vacant':
    #     # read teacherwise timetable and generate a vacant sheet
    #     # containing number of vacant periods for every teacher on each day.
    #     generate_vacant_sheet(book, context)
    #     book.save(args.infile)
    #     print(f"Vacant periods sheet saved to 'git {args.infile}'.")    

    elif args.command == 'diff':
        base = args.base
        current = args.current

        # compare "base" with "current"
        print(f"Comparing '{base}' with '{current}' ..." )
        differences = show_differences(base, current)
        print(f"Found {differences} differences between {base} and {current}.")
    else:
        print("twig.py -- timetable manipulation utility")
        print("Copyright (c) 2024 Sunil Sangwal <sunil.sangwal@gmail.com>")
        print("Type 'python twig.py -h' for more information.")
        exit(0)
        

    endTime = time.time()
    print("Finished processing in %.3f seconds." % (endTime - startTime))
    verbose("Have a nice day!\n", level=2)
    
    return warnings


if __name__ == '__main__':
    warnings = main()
    if warnings:
        exit(1)
    exit(0)
